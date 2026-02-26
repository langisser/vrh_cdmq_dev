#!/usr/bin/env python3
"""
run_insert_source_devtest.py

Reads source/Sample_Data_PoC_Match_Merge.xlsx and inserts data directly
into Databricks tables via DatabricksSession using multi-row INSERT VALUES.

Usage:
    source /home/khaw/ClaudeCode/databricks_dev_local/venv/bin/activate
    python3 scripts/run_insert_source_devtest.py
"""

import os
import sys
from pathlib import Path

os.environ["DATABRICKS_CONFIG_FILE"] = "/home/khaw/ClaudeCode/vrh_cdmq_dev/.databrickscfg"

import openpyxl
from databricks.connect import DatabricksSession

# ── Config ───────────────────────────────────────────────────────────────────
EXCEL_PATH  = Path("source/Sample_Data_PoC_Match_Merge.xlsx")
DATA_DT     = "2025-01-01"
MOTOR_TABLE = "viriyah_cdqm_poc.silver.source_motor_devtest"
TRUST_TABLE = "viriyah_cdqm_poc.silver.trust_source_devtest"
BATCH_SIZE  = 100

MOTOR_COLS = ["policy_id","id_card","fname","lname","gender","table",
              "birth_date","prefix","area","district","province",
              "postcode","email","phone_no","insert_date","update_date","DATA_DT"]

TRUST_COLS = ["id_card","fname","lname","gender","table",
              "birth_date","prefix","area","district","province",
              "postcode","email","phone_no","insert_date","update_date","DATA_DT"]


def escape(val) -> str:
    if val is None:
        return "NULL"
    s = str(val).strip()
    if s == "" or s.lower() == "none":
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def read_sheet(wb, sheet_name: str) -> list[dict]:
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        if not all(v is None for v in row.values()):
            rows.append(row)
    return rows


def build_col_list(cols: list[str]) -> str:
    return ", ".join(f"`{c}`" if c == "table" else c for c in cols)


def insert_batches(spark, table: str, cols: list[str], rows: list[dict]):
    col_list = build_col_list(cols)
    total = len(rows)
    batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

    for i in range(batches):
        batch = rows[i * BATCH_SIZE : (i + 1) * BATCH_SIZE]
        value_lines = []
        for row in batch:
            vals = []
            for col in cols:
                vals.append(f"'{DATA_DT}'" if col == "DATA_DT" else escape(row.get(col)))
            value_lines.append("  (" + ", ".join(vals) + ")")

        sql = f"INSERT INTO {table} ({col_list}) VALUES\n" + ",\n".join(value_lines)
        spark.sql(sql)
        print(f"  batch {i+1}/{batches} inserted ({len(batch)} rows)")


def main():
    spark = DatabricksSession.builder.getOrCreate()
    print(f"Spark connected: {spark.version}\n")

    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    motor_rows = read_sheet(wb, "Motor_source")
    trust_rows = read_sheet(wb, "Trust_source ")
    print(f"  Motor_source : {len(motor_rows)} rows")
    print(f"  Trust_source : {len(trust_rows)} rows\n")

    # ── Truncate ─────────────────────────────────────────────────────────────
    print("Truncating existing DATA_DT='2025-01-01' ...")
    spark.sql(f"DELETE FROM {MOTOR_TABLE} WHERE DATA_DT = '{DATA_DT}'")
    spark.sql(f"DELETE FROM {TRUST_TABLE} WHERE DATA_DT = '{DATA_DT}'")

    # ── Insert ────────────────────────────────────────────────────────────────
    print(f"\nInserting → {MOTOR_TABLE}")
    insert_batches(spark, MOTOR_TABLE, MOTOR_COLS, motor_rows)

    print(f"\nInserting → {TRUST_TABLE}")
    insert_batches(spark, TRUST_TABLE, TRUST_COLS, trust_rows)

    # ── Verify ────────────────────────────────────────────────────────────────
    print("\n── Verify ──")
    for tbl in [MOTOR_TABLE, TRUST_TABLE]:
        spark.sql(f"SELECT COUNT(*) AS cnt, DATA_DT FROM {tbl} GROUP BY DATA_DT").show()

    print("Done.")


if __name__ == "__main__":
    main()
