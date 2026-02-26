"""
Regenerate data_prep_dedup.py with properly padded timestamps.
Pads fractional seconds to 3 digits so to_timestamp('yyyy-MM-dd HH:mm:ss.SSS') works consistently.
"""
import openpyxl, re

def pad_ts(v):
    """Pad fractional seconds to 3 digits, return None if no value."""
    if v is None:
        return None
    s = str(v)
    m = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\.(\d+)$', s)
    if m:
        frac = m.group(2).ljust(3, '0')[:3]
        return f"{m.group(1)}.{frac}"
    return s

wb = openpyxl.load_workbook('/home/khaw/ClaudeCode/vrh_cdmq_dev/source/Sample_Data_PoC_Match_Merge.xlsx')

# ── Motor_source ──────────────────────────────────────────────
ws = wb['Motor_source']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
col_idx = {h: i+1 for i, h in enumerate(headers)}
src_cols = ['policy_id','id_card','fname','lname','gender','birth_date','prefix',
            'area','district','province','postcode','email','phone_no','insert_date','update_date']

source_rows = []
for r in range(2, ws.max_row+1):
    row = []
    for c in src_cols:
        v = ws.cell(r, col_idx[c]).value
        if c in ('insert_date', 'update_date'):
            row.append(pad_ts(v))
        elif v is None:
            row.append(None)
        else:
            row.append(str(v))
    source_rows.append(tuple(row))

# ── Trust_source ──────────────────────────────────────────────
ws2 = wb['Trust_source ']
headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column+1)]
col_idx2 = {h: i+1 for i, h in enumerate(headers2)}
ts_cols = ['id_card','fname','lname','gender','birth_date','prefix',
           'area','district','province','postcode','email','phone_no','insert_date','update_date']

trust_rows = []
for r in range(2, ws2.max_row+1):
    row = []
    for c in ts_cols:
        v = ws2.cell(r, col_idx2[c]).value
        if c in ('insert_date', 'update_date'):
            row.append(pad_ts(v))
        elif v is None:
            row.append(None)
        else:
            row.append(str(v))
    trust_rows.append(tuple(row))

print(f"source_rows: {len(source_rows)}")
print(f"trust_rows: {len(trust_rows)}")

# Check sample timestamps are padded
print("Sample insert_dates:")
for r in source_rows[:5]:
    print(f"  insert={r[13]!r}  update={r[14]!r}")

# ── Write notebook ────────────────────────────────────────────
def repr_tuple(t):
    parts = []
    for v in t:
        if v is None:
            parts.append('None')
        else:
            parts.append(repr(v))
    return '    (' + ', '.join(parts) + '),'

src_lines = [repr_tuple(r) for r in source_rows]
ts_lines  = [repr_tuple(r) for r in trust_rows]

header = '''# Databricks notebook source
# MAGIC %md
# MAGIC # Data Preparation — Dedup Unit Test
# MAGIC
# MAGIC Loads test data into devtest tables for the dedup unit test pipeline.
# MAGIC Data source: `Sample_Data_PoC_Match_Merge.xlsx` (Motor_source + Trust_source sheets)
# MAGIC
# MAGIC ## Tables Written
# MAGIC ### Silver (`viriyah_cdqm_poc.silver`)
# MAGIC - `source_motor_devtest` — {src_n} rows from Motor_source sheet, DATA_DT = '2026-02-26'
# MAGIC - `trust_source_devtest` — {ts_n} rows from Trust_source sheet, DATA_DT = '2026-02-26'
# MAGIC
# MAGIC ## Notes
# MAGIC - Column names match Excel headers exactly
# MAGIC - `insert_date` / `update_date` padded to 3 fractional digits, parsed as `yyyy-MM-dd HH:mm:ss.SSS`
# MAGIC - `id_card` in Trust_source cast to STRING
# MAGIC - Write mode: `overwrite` on partition `DATA_DT='2026-02-26'`

# COMMAND ----------

from pyspark.sql.types import StructType, StructField, StringType
from pyspark.sql.functions import lit, to_timestamp

catalog = 'viriyah_cdqm_poc'
silver  = 'silver'
DATA_DT = '2026-02-26'

print(f"catalog : {{catalog}}")
print(f"silver  : {{silver}}")
print(f"DATA_DT : {{DATA_DT}}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Step 0: Clean previous test run data (idempotent)

# COMMAND ----------

fw_schema   = 'control_fw'
bkey_schema = silver

prev_bkeys_rows = spark.sql(f"""
    SELECT DISTINCT bkey FROM {{catalog}}.{{bkey_schema}}.chv_table_bkey_v2
    WHERE lower(table) LIKE '%devtest%'
""").collect()
prev_bkeys = [r.bkey for r in prev_bkeys_rows]
print(f"Previous devtest BKEYs to clean from dedup tables: {{prev_bkeys}}")

if prev_bkeys:
    bkey_str = ", ".join(str(b) for b in prev_bkeys)
    for dedup_tbl in ["dedup_customer_name", "dedup_province", "dedup_gender", "dedup_email", "dedup_phone"]:
        spark.sql(f"DELETE FROM {{catalog}}.{{silver}}.{{dedup_tbl}} WHERE bkey IN ({{bkey_str}})")
        print(f"  Cleaned {{dedup_tbl}} (bkeys: {{bkey_str}})")

spark.sql(f"DELETE FROM {{catalog}}.{{bkey_schema}}.chv_table_bkey_v2 WHERE lower(table) LIKE '%devtest%'")
print("Cleaned chv_table_bkey_v2 (devtest entries)")

spark.sql(f"DELETE FROM {{catalog}}.{{fw_schema}}.chv_matching_result_v2 WHERE lower(MAIN_TABLE) LIKE '%devtest%'")
print("Cleaned chv_matching_result_v2 (devtest entries)")

spark.sql(f"DELETE FROM {{catalog}}.{{fw_schema}}.chv_matching_log_v2 WHERE lower(MAIN_TABLE) LIKE '%devtest%'")
print("Cleaned chv_matching_log_v2 (devtest entries)")

# COMMAND ----------

# MAGIC %md
# MAGIC ## source_motor_devtest — {src_n} rows from Motor_source sheet

# COMMAND ----------

src_schema = StructType([
    StructField("policy_id",   StringType(), True),
    StructField("id_card",     StringType(), True),
    StructField("fname",       StringType(), True),
    StructField("lname",       StringType(), True),
    StructField("gender",      StringType(), True),
    StructField("birth_date",  StringType(), True),
    StructField("prefix",      StringType(), True),
    StructField("area",        StringType(), True),
    StructField("district",    StringType(), True),
    StructField("province",    StringType(), True),
    StructField("postcode",    StringType(), True),
    StructField("email",       StringType(), True),
    StructField("phone_no",    StringType(), True),
    StructField("insert_date", StringType(), True),
    StructField("update_date", StringType(), True),
])

source_data = [
'''.format(src_n=len(source_rows), ts_n=len(trust_rows))

mid1 = ''']

src_df = spark.createDataFrame(source_data, schema=src_schema)
src_df = (src_df
    .withColumn("insert_date", to_timestamp("insert_date", "yyyy-MM-dd HH:mm:ss.SSS"))
    .withColumn("update_date", to_timestamp("update_date", "yyyy-MM-dd HH:mm:ss.SSS"))
    .withColumn("DATA_DT", lit(DATA_DT))
)

src_table = f"{catalog}.{silver}.source_motor_devtest"
src_df.write.format("delta").mode("overwrite") \\
    .option("replaceWhere", f"DATA_DT = '{DATA_DT}'") \\
    .partitionBy("DATA_DT") \\
    .saveAsTable(src_table)

src_cnt = src_df.count()
print(f"Written: {src_table} ({src_cnt} rows, DATA_DT={DATA_DT})")

# COMMAND ----------

# MAGIC %md
# MAGIC ## trust_source_devtest — {ts_n} rows from Trust_source sheet

# COMMAND ----------

ts_schema = StructType([
    StructField("id_card",     StringType(), True),
    StructField("fname",       StringType(), True),
    StructField("lname",       StringType(), True),
    StructField("gender",      StringType(), True),
    StructField("birth_date",  StringType(), True),
    StructField("prefix",      StringType(), True),
    StructField("area",        StringType(), True),
    StructField("district",    StringType(), True),
    StructField("province",    StringType(), True),
    StructField("postcode",    StringType(), True),
    StructField("email",       StringType(), True),
    StructField("phone_no",    StringType(), True),
    StructField("insert_date", StringType(), True),
    StructField("update_date", StringType(), True),
])

trust_data = [
'''.format(ts_n=len(trust_rows))

mid2 = ''']

ts_df = spark.createDataFrame(trust_data, schema=ts_schema)
ts_df = (ts_df
    .withColumn("insert_date", to_timestamp("insert_date", "yyyy-MM-dd HH:mm:ss.SSS"))
    .withColumn("update_date", to_timestamp("update_date", "yyyy-MM-dd HH:mm:ss.SSS"))
    .withColumn("DATA_DT", lit(DATA_DT))
)

ts_table = f"{catalog}.{silver}.trust_source_devtest"
ts_df.write.format("delta").mode("overwrite") \\
    .option("replaceWhere", f"DATA_DT = '{DATA_DT}'") \\
    .partitionBy("DATA_DT") \\
    .saveAsTable(ts_table)

ts_cnt = ts_df.count()
print(f"Written: {ts_table} ({ts_cnt} rows, DATA_DT={DATA_DT})")

# COMMAND ----------

# MAGIC %md
# MAGIC ## CHV_PRE_VALIDATION_RESULT_V2 — Seed PASSED rows

# COMMAND ----------

from pyspark.sql.types import LongType

PRCS_NM = 'TEST_PRE_VLD_DEVTEST'

src_vld_rows = []
for row in spark.sql(f"SELECT policy_id, id_card, fname, lname FROM {src_table} WHERE DATA_DT = '{DATA_DT}'").collect():
    key = row.policy_id or 'null_val'
    for col_name in ['id_card', 'fname', 'lname']:
        val = getattr(row, col_name) or 'N/A'
        result = 'PASSED' if getattr(row, col_name) else 'FAILED'
        src_vld_rows.append((src_table, key, 'CHECK_NULL', col_name, val, result, 1, PRCS_NM, 1))

ts_vld_rows = []
for row in spark.sql(f"SELECT id_card, fname, lname FROM {ts_table} WHERE DATA_DT = '{DATA_DT}'").collect():
    key = row.id_card or 'null_val'
    for col_name in ['id_card', 'fname', 'lname']:
        val = getattr(row, col_name) or 'N/A'
        result = 'PASSED' if getattr(row, col_name) else 'FAILED'
        ts_vld_rows.append((ts_table, key, 'CHECK_NULL', col_name, val, result, 1, PRCS_NM, 1))

vld_schema = StructType([
    StructField("TABLE",        StringType(), True),
    StructField("KEY",          StringType(), True),
    StructField("RULES",        StringType(), True),
    StructField("COLUMN",       StringType(), True),
    StructField("VALUE",        StringType(), True),
    StructField("RESULT",       StringType(), True),
    StructField("LD_ID",        LongType(),   True),
    StructField("UPDT_PRCS_NM", StringType(), True),
    StructField("UPDT_LD_ID",   LongType(),   True),
])

all_vld_rows = src_vld_rows + ts_vld_rows
vld_df = spark.createDataFrame(all_vld_rows, schema=vld_schema)

vld_tbl = f"{catalog}.{fw_schema}.CHV_PRE_VALIDATION_RESULT_V2"
vld_df.createOrReplaceTempView("vld_tmp")
spark.sql(f"""
    INSERT OVERWRITE {vld_tbl}
    PARTITION (DATA_DT = '{DATA_DT}', PRCS_NM = '{PRCS_NM}')
    SELECT TABLE, KEY, RULES, COLUMN, VALUE, RESULT, LD_ID, UPDT_PRCS_NM, UPDT_LD_ID
    FROM vld_tmp
""")

vld_cnt = spark.sql(f"SELECT COUNT(*) AS n FROM {vld_tbl} WHERE DATA_DT = '{DATA_DT}' AND PRCS_NM = '{PRCS_NM}'").collect()[0].n
print(f"Written: {vld_tbl} ({vld_cnt} pre-validation rows, DATA_DT={DATA_DT})")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Summary

# COMMAND ----------

print("=" * 60)
print("data_prep_dedup — Summary")
print("=" * 60)
print(f"  DATA_DT                : {DATA_DT}")
print(f"  source_motor_devtest   : {src_cnt} rows")
print(f"  trust_source_devtest   : {ts_cnt} rows")
print(f"  pre-validation results : {vld_cnt} rows")
print("=" * 60)
print("DONE")
'''

out_path = '/home/khaw/ClaudeCode/vrh_cdmq_dev/notebooks/work/unittest/dedup/data_prep_dedup.py'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(header)
    f.write('\n'.join(src_lines))
    f.write('\n')
    f.write(mid1)
    f.write('\n'.join(ts_lines))
    f.write('\n')
    f.write(mid2)

print(f"\nWritten: {out_path}")
import os
print(f"File size: {os.path.getsize(out_path):,} bytes")
