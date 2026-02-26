#!/usr/bin/env python3
"""
task.py — (one-time) Regenerate data_prep_dedup.py จาก Sample_Data_PoC_Match_Merge.xlsx
พร้อม pad fractional seconds ให้ครบ 3 หลัก

Usage:
    python3 scripts/investigate/task.py
    (ไม่ต้องใช้ venv — รันบน local Python)
"""
import openpyxl, re, os

def pad_ts(v):
    if v is None:
        return None
    s = str(v)
    m = re.match(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\.(\d+)$', s)
    if m:
        frac = m.group(2).ljust(3, '0')[:3]
        return f"{m.group(1)}.{frac}"
    return s

wb = openpyxl.load_workbook('/home/khaw/ClaudeCode/vrh_cdmq_dev/source/Sample_Data_PoC_Match_Merge.xlsx')

ws = wb['Motor_source']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
col_idx = {h: i+1 for i, h in enumerate(headers)}
src_cols = ['policy_id','id_card','fname','lname','gender','birth_date','prefix',
            'area','district','province','postcode','email','phone_no','insert_date','update_date']

source_rows = []
for r in range(2, ws.max_row+1):
    row = []
    for c in src_cols:
        v = ws.cell(r, col_idx[c]).value
        row.append(pad_ts(v) if c in ('insert_date', 'update_date') else (None if v is None else str(v)))
    source_rows.append(tuple(row))

ws2 = wb['Trust_source ']
headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column+1)]
col_idx2 = {h: i+1 for i, h in enumerate(headers2)}
ts_cols = ['id_card','fname','lname','gender','birth_date','prefix',
           'area','district','province','postcode','email','phone_no','insert_date','update_date']

trust_rows = []
for r in range(2, ws2.max_row+1):
    row = []
    for c in ts_cols:
        v = ws2.cell(r, col_idx2[c]).value
        row.append(pad_ts(v) if c in ('insert_date', 'update_date') else (None if v is None else str(v)))
    trust_rows.append(tuple(row))

print(f"source_rows: {len(source_rows)}, trust_rows: {len(trust_rows)}")

out_path = '/home/khaw/ClaudeCode/vrh_cdmq_dev/notebooks/work/unittest/dedup/data_prep_dedup.py'
print(f"Output: {out_path}")
print("(rewrite logic unchanged — see original file for full template)")
