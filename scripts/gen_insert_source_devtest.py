"""
gen_insert_source_devtest.py

Reads source/Sample_Data_PoC_Match_Merge.xlsx and generates INSERT INTO statements
for source_motor_devtest and trust_source_devtest tables.

Output: notebooks/work/match_and_merge/insert_scripts/insert_source_devtest.py

Usage:
    python3 scripts/gen_insert_source_devtest.py
"""

import openpyxl
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_PATH   = Path("source/Sample_Data_PoC_Match_Merge.xlsx")
OUTPUT_PATH  = Path("notebooks/work/match_and_merge/insert_scripts/insert_source_devtest.py")
DATA_DT      = "2025-01-01"
CATALOG      = "viriyah_cdqm_poc"
SILVER       = "silver"
MOTOR_TABLE  = f"{CATALOG}.{SILVER}.source_motor_devtest"
TRUST_TABLE  = f"{CATALOG}.{SILVER}.trust_source_devtest"
BATCH_SIZE   = 100   # rows per INSERT statement

# ── Column definitions ───────────────────────────────────────────────────────
MOTOR_COLS = [
    "policy_id", "id_card", "fname", "lname", "gender", "table",
    "birth_date", "prefix", "area", "district", "province",
    "postcode", "email", "phone_no", "insert_date", "update_date", "DATA_DT"
]

TRUST_COLS = [
    "id_card", "fname", "lname", "gender", "table",
    "birth_date", "prefix", "area", "district", "province",
    "postcode", "email", "phone_no", "insert_date", "update_date", "DATA_DT"
]


def escape(val) -> str:
    """Convert a Python value to SQL literal."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    if s == "" or s.lower() == "none":
        return "NULL"
    # Escape single quotes
    s = s.replace("'", "''")
    return f"'{s}'"


def read_sheet(wb, sheet_name: str, excel_cols: list[str]) -> list[dict]:
    """Read a sheet and return list of dicts (only excel_cols, not DATA_DT)."""
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        # Check if row is entirely empty
        if all(v is None for v in row_vals.values()):
            continue
        rows.append(row_vals)
    return rows


def build_insert_blocks(table_name: str, cols: list[str], rows: list[dict],
                        batch_size: int) -> list[str]:
    """Build batched INSERT INTO ... VALUES (...) blocks."""
    # cols includes DATA_DT at the end; excel_cols excludes DATA_DT
    col_list = ", ".join(f"`{c}`" if c == "table" else c for c in cols)
    blocks = []

    for start in range(0, len(rows), batch_size):
        batch = rows[start: start + batch_size]
        value_lines = []
        for row in batch:
            vals = []
            for col in cols:
                if col == "DATA_DT":
                    vals.append(f"'{DATA_DT}'")
                else:
                    vals.append(escape(row.get(col)))
            value_lines.append("  (" + ", ".join(vals) + ")")
        sql = (
            f"INSERT INTO {table_name} ({col_list}) VALUES\n"
            + ",\n".join(value_lines)
            + ";"
        )
        blocks.append(sql)

    return blocks


def to_notebook_cell(sql: str, comment: str | None = None) -> str:
    """Wrap SQL as a Databricks %sql magic cell."""
    header = f"# MAGIC %md\n# MAGIC {comment}\n\n# COMMAND ----------\n\n" if comment else ""
    lines = "\n".join(f"# MAGIC {line}" if line.strip() else "# MAGIC" for line in sql.split("\n"))
    return header + "# MAGIC %sql\n" + lines


def main():
    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    motor_rows = read_sheet(wb, "Motor_source", MOTOR_COLS)
    trust_rows = read_sheet(wb, "Trust_source ", TRUST_COLS)

    print(f"  Motor_source: {len(motor_rows)} data rows")
    print(f"  Trust_source: {len(trust_rows)} data rows")

    # Build SQL blocks
    motor_blocks = build_insert_blocks(MOTOR_TABLE, MOTOR_COLS, motor_rows, BATCH_SIZE)
    trust_blocks = build_insert_blocks(TRUST_TABLE, TRUST_COLS, trust_rows, BATCH_SIZE)

    # ── Assemble notebook ────────────────────────────────────────────────────
    sections = []
    sections.append("# Databricks notebook source")
    sections.append(
        "# MAGIC %md\n"
        "# MAGIC # INSERT — Source devtest data\n"
        "# MAGIC\n"
        f"# MAGIC Inserts data from `Sample_Data_PoC_Match_Merge.xlsx` into devtest tables.\n"
        "# MAGIC\n"
        "# MAGIC | Table | Rows | DATA_DT |\n"
        "# MAGIC |---|---|---|\n"
        f"# MAGIC | `{MOTOR_TABLE}` | {len(motor_rows)} | {DATA_DT} |\n"
        f"# MAGIC | `{TRUST_TABLE}` | {len(trust_rows)} | {DATA_DT} |\n"
        "# MAGIC\n"
        "# MAGIC Run **after** `ddl_source_devtest.py`."
    )

    # ── TRUNCATE section (idempotent) ────────────────────────────────────────
    sections.append(
        "# MAGIC %md\n"
        "# MAGIC ## Step 0: Truncate existing devtest data (idempotent re-run)"
    )
    truncate_sql = (
        f"DELETE FROM {MOTOR_TABLE} WHERE DATA_DT = '{DATA_DT}';\n\n"
        f"DELETE FROM {TRUST_TABLE} WHERE DATA_DT = '{DATA_DT}';"
    )
    sections.append(to_notebook_cell(truncate_sql))

    # ── Motor source inserts ─────────────────────────────────────────────────
    sections.append(
        "# MAGIC %md\n"
        f"# MAGIC ## Step 1: Insert source_motor_devtest ({len(motor_rows)} rows, "
        f"{len(motor_blocks)} batches × {BATCH_SIZE})"
    )
    for i, block in enumerate(motor_blocks, 1):
        comment = f"### Motor batch {i}/{len(motor_blocks)}" if len(motor_blocks) > 1 else None
        sections.append(to_notebook_cell(block, comment))

    # ── Trust source inserts ─────────────────────────────────────────────────
    sections.append(
        "# MAGIC %md\n"
        f"# MAGIC ## Step 2: Insert trust_source_devtest ({len(trust_rows)} rows, "
        f"{len(trust_blocks)} batches × {BATCH_SIZE})"
    )
    for i, block in enumerate(trust_blocks, 1):
        comment = f"### Trust batch {i}/{len(trust_blocks)}" if len(trust_blocks) > 1 else None
        sections.append(to_notebook_cell(block, comment))

    # ── Verify ───────────────────────────────────────────────────────────────
    sections.append(
        "# MAGIC %md\n"
        "# MAGIC ## Verify"
    )
    verify_sql = (
        f"SELECT COUNT(*) AS cnt, DATA_DT FROM {MOTOR_TABLE} GROUP BY DATA_DT ORDER BY DATA_DT;\n\n"
        f"SELECT COUNT(*) AS cnt, DATA_DT FROM {TRUST_TABLE} GROUP BY DATA_DT ORDER BY DATA_DT;"
    )
    sections.append(to_notebook_cell(verify_sql))

    # ── Write output ─────────────────────────────────────────────────────────
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    content = "\n\n# COMMAND ----------\n\n".join(sections) + "\n"
    OUTPUT_PATH.write_text(content, encoding="utf-8")

    print(f"\nOutput written: {OUTPUT_PATH}")
    print(f"  Motor blocks : {len(motor_blocks)}")
    print(f"  Trust blocks : {len(trust_blocks)}")
    print("Done.")


if __name__ == "__main__":
    main()
